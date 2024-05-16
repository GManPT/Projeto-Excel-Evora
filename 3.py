import openpyxl

workbook = openpyxl.load_workbook("alunos.xlsx")
folha_turma = workbook["Turma A"]

# Procura o nome "Alan Elton Ramos"
linha_alan = 0
for row in range(2, folha_turma.max_row + 1):
    if folha_turma.cell(row=row, column=2).value == "Alan Elton Ramos":
        linha_alan = row
        break

# Muda o valor da celula de matematica
folha_turma.cell(row=linha_alan, column=6, value=8.93345)

# Muda o valor da celula de historia
folha_turma.cell(row=linha_alan, column=8, value=7.334556)

# Guardar as alteracoes
workbook.save("alunos.xlsx")
