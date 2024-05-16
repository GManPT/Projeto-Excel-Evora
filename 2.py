import openpyxl
from copy import copy
from datetime import datetime

def calcula_idade(data_nascimento):
    data_atual = datetime.now()

    # Converter a string para um objeto datetime
    data_nascimento = datetime.strptime(data_nascimento, "%Y-%m-%d %H:%M:%S")
    idade = data_atual.year - data_nascimento.year

    # Verificar se o aniversário já ocorreu este ano
    if (data_atual.month, data_atual.day) < (data_nascimento.month, data_nascimento.day):
        idade -= 1
    
    return idade

workbook = openpyxl.load_workbook("alunos.xlsx")
folha_turma = workbook["Turma A"]

# Inserir a coluna da "idade" imediatamente após a data de nascimento
folha_turma.insert_cols(14)
celula_anterior = folha_turma.cell(row=1, column=4)
celula_titulo = folha_turma.cell(row=1, column=14, value="Idade")

# Copiar formatacao
celula_titulo.font = copy(celula_anterior.font)
celula_titulo.border = copy(celula_anterior.border)
celula_titulo.fill = copy(celula_anterior.fill)
celula_titulo.alignment = copy(celula_anterior.alignment)

# Inserir as celulas com as idades
for row in range(2, folha_turma.max_row + 1):
    cell_data = folha_turma.cell(row=row, column=4)
    idade = calcula_idade(str(cell_data.value))
    folha_turma.cell(row=row, column=14, value=idade)

# Inserir se o aluno foi aprovado ou nao
for row in range(2, folha_turma.max_row + 1):
    soma = 0
    for coluna in ['F', 'G', 'H']:
        valor_celula = folha_turma[f"{coluna}{row}"].value
        soma += valor_celula

    media_final = soma / 3

    if media_final >= 5:
        folha_turma.cell(row=row, column=13, value="Aprovado")
    else:
        folha_turma.cell(row=row, column=13, value="Reprovado")

# Guardar as alteracoes
workbook.save("alunos.xlsx")


