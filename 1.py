import openpyxl
from copy import copy

# Carregar o arquivo Excel
workbook = openpyxl.load_workbook("alunos.xlsx")

# Remover a folha com o nome "Folha1"
workbook.remove(workbook["Folha1"])
folha_original = workbook["Turma A"]
nova_folha = workbook.create_sheet(title=f'Cópia do Original')

# Copiar o conteúdo da folha original para a nova folha
for row in folha_original.iter_rows():
    for cell in row:
        new_cell = nova_folha.cell(row=cell.row, column=cell.col_idx, value=cell.value)
        # Copiar estilo da celula
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.alignment = copy(cell.alignment)

# Copiar dimensões das colunas
for col in folha_original.column_dimensions:
    nova_folha.column_dimensions[col] = copy(folha_original.column_dimensions[col])

# Copiar dimensões das linhas
for row_dim in folha_original.row_dimensions:
    nova_folha.row_dimensions[row_dim] = copy(folha_original.row_dimensions[row_dim])

# Guardar as alteracoes
workbook.save("alunos.xlsx")